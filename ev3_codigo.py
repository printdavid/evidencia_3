import datetime
import sys
import openpyxl
from openpyxl.styles import Font, Alignment
import sqlite3
from sqlite3 import Error
import os.path

file_exists = os.path.exists("registros.db")

if file_exists:
  print("ARCHIVO DB ENCONTRADO")
else:
  print("ARCHIVO DB NO ENCONTRADO, PRIMERA EJECUCION ASUMIDA")
  try:
    with sqlite3.connect("registros.db") as conn:
      mi_cursor = conn.cursor()
      mi_cursor.execute("CREATE TABLE IF NOT EXISTS sala (clave_sala INTEGER PRIMARY KEY, nombre TEXT NOT NULL, cupo INTEGER NOT NULL);")
      mi_cursor.execute("CREATE TABLE IF NOT EXISTS cliente (clave_cliente INTEGER PRIMARY KEY, nombre TEXT NOT NULL);")
      mi_cursor.execute("CREATE TABLE IF NOT EXISTS turno (clave_turno INTEGER PRIMARY KEY, nombre TEXT NOT NULL);")
      mi_cursor.execute("CREATE TABLE IF NOT EXISTS reservacion (clave_reservacion INTEGER PRIMARY KEY, nombre TEXT NOT NULL, fecha TIMESTAMP NOT NULL, "
                        "clave_sala INTEGER NOT NULL, clave_cliente INTEGER NOT NULL, clave_turno INTEGER NOT NULL, "
                        "FOREIGN KEY(clave_sala) REFERENCES sala(clave_sala), FOREIGN KEY(clave_cliente) REFERENCES cliente(clave_cliente), FOREIGN KEY(clave_turno) REFERENCES turno(clave_turno));")

      valores = {"clave_turno":1, "nombre":"MATUTINO"}
      mi_cursor.execute("INSERT INTO turno VALUES(:clave_turno,:nombre)", valores)
      valores = {"clave_turno":2, "nombre":"VESPERTINO"}
      mi_cursor.execute("INSERT INTO turno VALUES(:clave_turno,:nombre)", valores)
      valores = {"clave_turno":3, "nombre":"NOCTURNO"}
      mi_cursor.execute("INSERT INTO turno VALUES(:clave_turno,:nombre)", valores)
  except Error as e:
    print(e)
  except:
    print(f"Se produjo el siguiente error: {sys.exc_info()[0]}\n")
  finally:
    conn.close()

while True:
  print(f'\n{"-" * 70}')
  print(f'{" " * 23}    MENU DE OPCIONES    {" " * 23}')
  print(f'{"-" * 70}')
  print(f"\n [1] Submenu - Reservaciones\n",
        "[2] Submenu - Reportes\n",
        "[3] Registrar una Sala\n",
        "[4] Registrar un Cliente\n",
        "[5] Salir")
  print(f"{'~' * 70}")

  while True:
    try:
      opcion_menu = int(input("SELECCIONE UNA OPCION: "))
    except ValueError:
      print("INGRESE UN VALOR NUMERICO\n")
    except Exception:
      print(f"OCURRIO UN PROBLEMA {sys.exc_info()[0]}\n")
    else:
      if opcion_menu in [1,2,3,4,5]:
        break
      else:
        print("INGRESE UNA RESPUESTA VALIDA\n")

  if opcion_menu == 1:
    while True:
      print(f'\n{"-" * 70}')
      print(f'{" " * 23}    SUBMENU DE RESERVACIONES    {" " * 23}')
      print(f'{"-" * 70}')
      print(f"\n [1] Registrar una Reservacion\n",
            "[2] Editar el Nombre de un Evento ya Existente\n",
            "[3] Consultar Disponibilidad de Salas para una Fecha\n",
            "[4] Eliminar una Reservacion\n",
            "[5] Volver al Menu Principal\n")
      print(f"{'~' * 70}")

      while True:
        try:
          opcion_reservaciones = int(input("SELECCIONE UNA OPCION: "))
        except ValueError:
          print("INGRESE UN VALOR NUMERICO\n")
        except Exception:
          print(f"OCURRIO UN PROBLEMA {sys.exc_info()[0]}\n")
        else:
          if opcion_reservaciones in [1,2,3,4,5]:
            break
          else:
            print("INGRESE UNA RESPUESTA VALIDA\n")

      if opcion_reservaciones == 1:
        try:
          with sqlite3.connect("registros.db") as conn:
            mi_cursor = conn.cursor()
            mi_cursor.execute("SELECT * FROM cliente")
            registros_cliente = mi_cursor.fetchall()

            mi_cursor.execute("SELECT * FROM sala")
            registros_sala = mi_cursor.fetchall()
        except Error as e:
          print (e)
        except Exception:
          print(f"Se produjo el siguiente error: {sys.exc_info()[0]}\n")
        finally:
          conn.close()

        if registros_sala and registros_cliente:
          print(f'\n{"_" * 70}')
          print(f'{" " * 20}  MENU REGISTRO DE UNA RESERVACION  {" " * 20}')
          print(f'{"_" * 70}')

          print("\n****** CLIENTES EXISTENTES ******")
          print("*" * 28)
          print("{:<6} {:<20}".format('CLAVE','NOMBRE'))
          print("*" * 28)

          for clave_cliente, nombre in registros_cliente:
            print("{:<6} {:<20}".format(clave_cliente, nombre))
          print(f'{"*" * 28}\n')

          claves_cliente = [clave_cliente for clave_cliente, nombre in registros_cliente]

          while True:
            try:
              clave_cliente = int(input("INGRESE LA CLAVE DEL CLIENTE: "))
            except ValueError:
              print("INGRESE UN VALOR NUMERICO\n")
            except Exception:
              print(f"OCURRIO UN PROBLEMA {sys.exc_info()[0]}\n")
            else:
              if clave_cliente in claves_cliente:
                break
              else:
                print("INGRESE UNA CLAVE DE CLIENTE EXISTENTE\n")

          while True:
            try:
              fecha_capturada = input("INGRESE LA FECHA DEL EVENTO (DD/MM/AAAA): ")
              fecha_reservacion = datetime.datetime.strptime(fecha_capturada, "%d/%m/%Y").date()
              fecha_datetime = datetime.datetime(fecha_reservacion.year, fecha_reservacion.month, fecha_reservacion.day)
            except ValueError:
              print("INGRESE UNA FECHA POSIBLE CON EL FORMATO CORRESPONDIENTE\n")
            except Exception:
              print(f"OCURRIO UN PROBLEMA {sys.exc_info()[0]}\n")
            else:
              fecha_actual = datetime.date.today()
              fecha_valida = fecha_reservacion - datetime.timedelta(days=2)

              if fecha_actual <= fecha_valida:
                break
              else:
                print("LA RESERVACION DEBE DE HACERSE CON 2 DIAS DE ANTICPACION ANTES DEL EVENTO\n")

          claves_sala = [clave_sala for clave_sala, nombre, cupo in registros_sala]

          print("\n****** SALAS EXISTENTES ******")
          print("*" * 28)
          print("{:<6} {:<20}".format('CLAVE','NOMBRE'))
          print("*" * 28)
          for clave_sala, nombre, cupo in registros_sala:
            print("{:<6} {:<20}".format(clave_sala, nombre))
          print(f'{"*" * 28}\n')

          while True:
            try:
              clave_sala = int(input("INGRESE LA CLAVE DE LA SALA: "))
            except ValueError:
              print("INGRESE UN VALOR NUMERICO\n")
            except Exception:
              print(f"OCURRIO UN PROBLEMA {sys.exc_info()[0]}\n")
            else:
              if clave_sala in claves_sala:
                break
              else:
                print("SE TIENE QUE ELEGIR UNA SALA EXISTENTE\n")

          print(f'{"~" * 70}')
          print(f'{" " * 20}   TURNOS DE RESERVACIONES   {" " * 20}')
          print(f" 1) MATUTINO\n",
                  "2) VESPERTINO\n",
                  "3) NOCTURNO\n")

          while True:
            try:
              clave_turno = int(input("SELECCIONE EL TURNO DE LA RESERVACION: "))
            except ValueError:
              print("INGRESE UN VALOR NUMERICO\n")
            except Exception:
              print(f"OCURRIO UN PROBLEMA {sys.exc_info()[0]}\n")
            else:
              if clave_turno in [1,2,3]:
                break
              else:
                print("INGRESE UNA OPCION VALIDA\n")

          try:
            with sqlite3.connect("registros.db", detect_types = sqlite3.PARSE_DECLTYPES | sqlite3.PARSE_COLNAMES) as conn:
              mi_cursor = conn.cursor()
              valores = {"clave_turno":clave_turno,"fecha_reservacion":fecha_datetime, "clave_sala":clave_sala}
              mi_cursor.execute("SELECT COUNT(*) FROM reservacion WHERE (clave_turno = :clave_turno) AND (fecha = :fecha_reservacion) AND (clave_sala = :clave_sala);", valores)
              registros = mi_cursor.fetchall()
              coincidencias = registros[0][0]
          except Error as e:
            print (e)
          except Exception:
            print(f"Se produjo el siguiente error: {sys.exc_info()[0]}\n")
          finally:
            conn.close()

          if coincidencias == 0:
            while True:
              try:
                nombre_reservacion = input("INGRESE EL NOMBRE DEL EVENTO: ")
              except Exception:
                print(f"OCURRIO UN PROBLEMA {sys.exc_info()[0]}\n")
              else:
                if nombre_reservacion == "":
                  print("EL DATO NO SE PUEDE OMITIR\n")
                else:
                  break

            try:
              with sqlite3.connect("registros.db", detect_types = sqlite3.PARSE_DECLTYPES | sqlite3.PARSE_COLNAMES) as conn:
                mi_cursor = conn.cursor()
                valores = {"nombre":nombre_reservacion,"fecha":fecha_datetime,"clave_sala":clave_sala,"clave_cliente":clave_cliente,"clave_turno":clave_turno}
                mi_cursor.execute("INSERT INTO reservacion (nombre,fecha,clave_sala,clave_cliente,clave_turno) VALUES(:nombre,:fecha,:clave_sala,:clave_cliente,:clave_turno);", valores)
                mi_cursor.execute("SELECT fecha FROM reservacion;")
                registros = mi_cursor.fetchall()
            except Error as e:
              print (e)
            except:
              print(f"Se produjo el siguiente error: {sys.exc_info()[0]}")
              print("HOLA")
            finally:
              conn.close()

            print(f'\n{"*" * 10} RESERVACION REGISTRADA CON EXITO {"*" * 10}')

          else:
            print("YA EXISTE UNA RESERVACION EN LA FECHA, TURNO y SALA INGRESADA\n")
        else:
          print("NO SE CUENTA CON CLIENTES O SALAS REGISTRADAS\n")

      elif opcion_reservaciones == 2:
        print(f'\n{"_" * 70}')
        print(f'{" " * 20}  EDITAR NOMBRE DEL EVENTO  {" " * 20}\n')
        print(f'{"_" * 70}')

        print("\n****** EVENTOS EXISTENTES ******")
        print("*" * 28)
        print("{:<6} {:<20}".format('CLAVE','NOMBRE'))
        print("*" * 28)

        try:
          with sqlite3.connect("registros.db") as conn:
            mi_cursor = conn.cursor()
            mi_cursor.execute("SELECT clave_reservacion, nombre FROM reservacion;")
            registros = mi_cursor.fetchall()

            if registros:
              for clave_reservacion, nombre in registros:
                print("{:<6} {:<20}".format(clave_reservacion, nombre))
              print(f'{"*" * 28}\n')
        except Error as e:
          print (e)
        except Exception:
          print(f"Se produjo el siguiente error: {sys.exc_info()[0]}\n")
        finally:
          conn.close()

        claves_reservacion = [clave_reservacion for clave_reservacion, nombre in registros]

        while True:
          try:
            clave = int(input("INGRESE LA CLAVE DE LA RESERVACION A MODIFICAR: "))
          except ValueError:
            print("INGRESE UN VALOR NUMERICO\n")
          except Exception:
            print(f"OCURRIO UN PROBLEMA {sys.exc_info()[0]}\n")
          else:
            if clave in claves_reservacion:
              break
            else:
              print("INGRESE UNA CLAVE DE RESERVACION EXISTENTE\n")

        while True:
          try:
            nombre_nuevo = input("INGRESE EL NUEVO NOMBRE DE LA SALA: ")
          except Exception:
            print(f"OCURRIO UN PROBLEMA {sys.exc_info()[0]}\n")
          else:
            if nombre_nuevo == "":
              print("EL DATO NO SE PUEDE OMITIR\n")
            else:
              break

        try:
          with sqlite3.connect("registros.db") as conn:
            mi_cursor = conn.cursor()
            valores = {"clave":clave, "nombre":nombre_nuevo}
            mi_cursor.execute("UPDATE reservacion SET nombre = :nombre WHERE clave_reservacion = :clave;", valores)
        except Error as e:
          print (e)
        except Exception:
          print(f"Se produjo el siguiente error: {sys.exc_info()[0]}\n")
        finally:
          conn.close()

        print(f"\n***** MODIFICACION EXITOSA *****")

      elif opcion_reservaciones == 3:
        while True:
          try:
            fecha_capturada = input("INGRESE LA FECHA A CONSULTAR (DD/MM/AAAA): ")
            fecha_modificada = datetime.datetime.strptime(fecha_capturada, "%d/%m/%Y").date()
          except ValueError:
            print("INGRESE UNA FECHA POSIBLE CON EL FORMATO CORRESPONDIENTE\n")
          except Exception:
            print(f"OCURRIO UN PROBLEMA {sys.exc_info()[0]}\n")
          else:
            break

        print(f'\n{"*" * 28}')
        print("****** DISPONIBILIDAD ******")
        print("*" * 28)
        print("{:<6} {:<10} {:<20}".format('CLAVE','SALA','TURNO'))
        print("*" * 28)

        try:
          with sqlite3.connect("registros.db") as conn:
            mi_cursor = conn.cursor()
            valores = {"fecha_modificada":fecha_modificada}
            mi_cursor.execute("SELECT s.clave_sala, s.nombre, t.nombre FROM sala s, turno t, reservacion r "
                              "WHERE r.clave_sala=s.clave_sala AND r.clave_turno=t.clave_turno "
                              "AND DATE(r.fecha)= :fecha_modificada;",valores)
            registros_hechos = mi_cursor.fetchall()

            mi_cursor.execute("SELECT clave_sala, nombre FROM sala;")
            salas_posibles = mi_cursor.fetchall()

            mi_cursor.execute("SELECT nombre FROM turno;")
            turnos_posibles = mi_cursor.fetchall()

            registros_posibles = []
            for clave_sala, nombre_sala in salas_posibles:
              for nombre_turno in turnos_posibles:
                registros_posibles.append((clave_sala, nombre_sala, nombre_turno[0]))

            if registros_hechos:
              reservaciones_registradas = set(registros_hechos)
              reservaciones_posibles = set(registros_posibles)
              reservaciones_disponibles = sorted(list(reservaciones_posibles - reservaciones_registradas))

              for clave_sala, nombre_sala, nombre_turno in reservaciones_disponibles:
                print("{:<6} {:<10} {:<20}".format(clave_sala, nombre_sala, nombre_turno))
            else:
              for clave_sala, nombre_sala, nombre_turno in reservaciones_posibles:
                print("{:<6} {:<10} {:<20}".format(clave_sala, nombre_sala, nombre_turno))
            print("*" * 28)
        except Error as e:
          print (e)
        except Exception:
          print(f"Se produjo el siguiente error: {sys.exc_info()[0]}\n")
        finally:
          conn.close()

      elif opcion_reservaciones == 4:
        print(f'\n{"_" * 70}')
        print(f'{" " * 20}  ELIMINACION DE RESERVACION  {" " * 20}\n')
        print(f'{"_" * 70}')

        print("\n****** EVENTOS EXISTENTES ******")
        print("*" * 28)
        print("{:<6} {:<20}".format('CLAVE','NOMBRE'))
        print("*" * 28)

        try:
          with sqlite3.connect("registros.db") as conn:
            mi_cursor = conn.cursor()
            mi_cursor.execute("SELECT clave_reservacion, nombre FROM reservacion;")
            registros = mi_cursor.fetchall()

            if registros:
              for clave_reservacion, nombre in registros:
                print("{:<6} {:<20}".format(clave_reservacion, nombre))
              print(f'{"*" * 28}\n')
        except Error as e:
          print (e)
        except Exception:
          print(f"Se produjo el siguiente error: {sys.exc_info()[0]}\n")
        finally:
          conn.close()

        claves_reservacion = [clave_reservacion for clave_reservacion, nombre in registros]

        while True:
          try:
            clave = int(input("INGRESE LA CLAVE DE LA RESERVACION: "))
          except ValueError:
            print("INGRESE UN VALOR NUMERICO\n")
          except Exception:
            print(f"OCURRIO UN PROBLEMA {sys.exc_info()[0]}\n")
          else:
            if clave in claves_reservacion:
              break
            else:
              print("INGRESE UNA CLAVE DE RESERVACION EXISTENTE\n")

        try:
          with sqlite3.connect("registros.db", detect_types = sqlite3.PARSE_DECLTYPES | sqlite3.PARSE_COLNAMES) as conn:
            mi_cursor = conn.cursor()
            valores = {"clave":clave}
            mi_cursor.execute("SELECT clave_reservacion, fecha FROM reservacion WHERE clave_reservacion = :clave",valores)
            registros = mi_cursor.fetchall()

            for clave_reservacion, fecha in registros:
              fecha_actual = datetime.date.today()
              fecha_actual = datetime.datetime(fecha_actual.year,fecha_actual.month,fecha_actual.day)
              fecha_valida = fecha - datetime.timedelta(days=3)
        except Error as e:
          print (e)
        except Exception:
          print(f"Se produjo el siguiente error: {sys.exc_info()[0]}\n")
        finally:
          conn.close()

        if fecha_actual <= fecha_valida:
          try:
            with sqlite3.connect("registros.db", detect_types = sqlite3.PARSE_DECLTYPES | sqlite3.PARSE_COLNAMES) as conn:
              mi_cursor = conn.cursor()
              valores = {"clave":clave}
              mi_cursor.execute("SELECT r.clave_reservacion, r.nombre, r.fecha, s.nombre, c.nombre, t.nombre FROM reservacion r, sala s, cliente c, turno t "
                                "WHERE r.clave_sala=s.clave_sala AND r.clave_cliente=c.clave_cliente AND r.clave_turno=t.clave_turno "
                                "AND r.clave_reservacion = :clave",valores)
              registros = mi_cursor.fetchall()

              for clave_reservacion, nombre, fecha, nombre_sala, nombre_cliente, nombre_turno in registros:
                print("\n****** CONFIRMACION ******\n")
                print("DATOS DE RESERVACION")
                print(f"CLAVE: {clave_reservacion}")
                print(f"NOMBRE: {nombre}")
                print(f"FECHA: {fecha.date().strftime('%d/%m/%Y')}")
                print(f"SALA: {nombre_sala}")
                print(f"CLIENTE: {nombre_cliente}")
                print(f"TURNO: {nombre_turno}\n")
          except Error as e:
            print (e)
          except Exception:
            print(f"Se produjo el siguiente error: {sys.exc_info()[0]}\n")
          finally:
            conn.close()

          while True:
            try:
              eleccion = input("¿CONFIRMAR ELIMINACION?, ESTE ES UN PROCEDIMIENTO QUE NO PUEDE SER DESHECHO [S/N]: ")
            except Exception:
              print(f"OCURRIO UN PROBLEMA {sys.exc_info()[0]}\n")
            else:
              if eleccion.upper() in ["S","N"]:
                break
              else:
                print("INGRESE UNA OPCION VALIDA\n")

          if eleccion.upper() == "S":
            try:
              with sqlite3.connect("registros.db", detect_types = sqlite3.PARSE_DECLTYPES | sqlite3.PARSE_COLNAMES) as conn:
                mi_cursor = conn.cursor()
                valores = {"clave":clave}
                mi_cursor.execute("DELETE FROM reservacion WHERE clave_reservacion = :clave;",valores)

                print(f"\n***** ELIMINACION EXITOSA *****")
            except Error as e:
              print (e)
            except Exception:
              print(f"Se produjo el siguiente error: {sys.exc_info()[0]}\n")
            finally:
              conn.close()

          elif eleccion.upper() == "N":
            print(f"***** ELIMINACION NO HECHA *****")
        else:
          print("SOLAMENTE SE PUEDEN ELIMINAR RESERVACIONES CON TRES DIAS DE ANTICIPACION")

      elif opcion_reservaciones == 5:
        break

  elif opcion_menu == 2:
    while True:
      print(f'\n{"-" * 70}')
      print(f'{" " * 23}    SUBMENU DE REPORTES    {" " * 23}')
      print(f'{"-" * 70}')
      print(f"\n [1] Reporte en Pantalla de Reservaciones para una Fecha\n",
            "[2] Exportar Reporte en Excel\n",
            "[3] Volver al Menu Principal\n")
      print(f"{'~' * 70}")

      while True:
        try:
          opcion_reportes = int(input("SELECCIONE UNA OPCION: "))
        except ValueError:
          print("INGRESE UN VALOR NUMERICO\n")
        except Exception:
          print(f"OCURRIO UN PROBLEMA {sys.exc_info()[0]}\n")
        else:
          if opcion_reportes in [1,2,3]:
            break
          else:
            print("INGRESE UNA OPCION VALIDA\n")

      if opcion_reportes == 1:
        print(f'\n{"_" * 70}')
        print(f'{" " * 20}  REPORTE DE RESERVACIONES  {" " * 20}')
        print(f'{"_" * 70}')

        while True:
          try:
            fecha_capturada = input("INGRESE LA FECHA DEL EVENTO A CONSULTAR  (DD/MM/AAAA): ")
            fecha_modificada = datetime.datetime.strptime(fecha_capturada, "%d/%m/%Y").date()
          except ValueError:
            print("INGRESE UNA FECHA POSIBLE CON EL FORMATO CORRESPONDIENTE\n")
          except Exception:
            print(f"OCURRIO UN PROBLEMA {sys.exc_info()[0]}\n")
          else:
            break

        print("\n" + "*"*77)
        print("**" + " "*13 + f"REPORTE DE RESERVACIONES PARA EL DÍA {fecha_capturada}" + " "*13 + "**")
        print("*"*77)
        print("{:<6} {:<20} {:<38} {:<13}".format('SALA','CLIENTE','EVENTO', 'TURNO'))
        print("*" * 77)

        try:
          with sqlite3.connect("registros.db") as conn:
            mi_cursor = conn.cursor()
            valores = {"fecha_modificada":fecha_modificada}
            mi_cursor.execute("SELECT s.nombre, c.nombre, r.nombre, t.nombre FROM sala s, cliente c, turno t, reservacion r "
                              "WHERE r.clave_sala=s.clave_sala AND r.clave_cliente=c.clave_cliente AND r.clave_turno=t.clave_turno "
                              "AND DATE(r.fecha)= :fecha_modificada",valores)
            registros = mi_cursor.fetchall()

            if registros:
              for nombre_sala, nombre_cliente, nombre_reservacion, nombre_turno in registros:
                print("{:<6} {:<20} {:<38} {:<13}".format(nombre_sala, nombre_cliente, nombre_reservacion, nombre_turno))
              print("*"*30 + " FIN DEL REPORTE " + "*"*30)
            else:
              print("*"*30 + " FIN DEL REPORTE " + "*"*30)
        except Error as e:
          print (e)
        except Exception:
          print(f"Se produjo el siguiente error: {sys.exc_info()[0]}\n")
        finally:
          conn.close()

      elif opcion_reportes == 2:
        print(f'\n{"_" * 70}')
        print(f'{" " * 20}  REPORTE DE RESERVACIONES EN EXCEL {" " * 20}')
        print(f'{"_" * 70}')

        while True:
          try:
            fecha_capturada = input("INGRESE LA FECHA DEL EVENTO A CONSULTAR  (DD/MM/AAAA): ")
            fecha_modificada = datetime.datetime.strptime(fecha_capturada, "%d/%m/%Y").date()
          except ValueError:
            print("INGRESE UNA FECHA POSIBLE CON EL FORMATO CORRESPONDIENTE\n")
          except Exception:
            print(f"OCURRIO UN PROBLEMA {sys.exc_info()[0]}\n")
          else:
            break

        try:
          with sqlite3.connect("registros.db") as conn:
            mi_cursor = conn.cursor()
            valores = {"fecha_modificada":fecha_modificada}
            mi_cursor.execute("SELECT s.nombre, c.nombre, r.nombre, t.nombre FROM sala s, cliente c, turno t, reservacion r "
                              "WHERE r.clave_sala=s.clave_sala AND r.clave_cliente=c.clave_cliente AND r.clave_turno=t.clave_turno "
                              "AND DATE(r.fecha)= :fecha_modificada",valores)
            registros = mi_cursor.fetchall()
        except Error as e:
          print (e)
        except Exception:
          print(f"Se produjo el siguiente error: {sys.exc_info()[0]}\n")
        finally:
          conn.close()

        libro = openpyxl.Workbook()
        libro.iso_dates = True
        hoja = libro["Sheet"]
        hoja.title = "Reservaciones"

        hoja.sheet_properties.tabColor = 'A4C5DF'

        hoja["A1"].value = f"REPORTE DE RESERVACIONES PARA EL DÍA {fecha_capturada}"

        hoja.merge_cells('A1:H1')
        hoja.merge_cells('A2:B2')
        hoja.merge_cells('C2:D2')
        hoja.merge_cells('E2:F2')
        hoja.merge_cells('G2:H2')

        hoja['A1'].font = Font(bold=True, italic=True, size = 14)
        hoja['A1'].alignment = Alignment (horizontal ='center')
        hoja['A2'].alignment = Alignment (horizontal='center')
        hoja['C2'].alignment = Alignment (horizontal='center')
        hoja['E2'].alignment = Alignment (horizontal='center')
        hoja['G2'].alignment = Alignment (horizontal='center')

        hoja["A2"].value = f"Sala"
        hoja["C2"].value = f"Cliente"
        hoja["E2"].value = f"Nombre"
        hoja["G2"].value = f"Turno"

        renglon = 3
        for nombre_sala, nombre_cliente, nombre_reservacion, nombre_turno in registros:
          hoja.merge_cells(start_row = renglon, start_column = 1, end_row = renglon, end_column = 2)
          hoja.merge_cells(start_row = renglon, start_column = 3, end_row = renglon, end_column = 4)
          hoja.merge_cells(start_row = renglon, start_column = 5, end_row = renglon, end_column = 6)
          hoja.merge_cells(start_row = renglon, start_column = 7, end_row = renglon, end_column = 8)

          hoja.cell(row = renglon, column = 1).value = nombre_sala
          hoja.cell(row = renglon, column = 3).value = nombre_cliente
          hoja.cell(row = renglon, column = 5).value = nombre_reservacion
          hoja.cell(row = renglon, column = 7).value = nombre_turno
          renglon += 1

        dir = "./reporte_excel.xlsx"
        libro.save(dir)

        print(f"***** REPORTE CREADO *****")

      elif opcion_reportes == 3:
        break

  elif opcion_menu == 3:
    print(f'\n{"_" * 70}')
    print(f'{" " * 20}  REGISTRO DE UNA SALA  {" " * 20}')
    print(f'{"_" * 70}')

    while True:
      try:
        nombre_sala = input("INGRESE EL NOMBRE DE LA SALA A REGISTRAR: ")
      except Exception:
        print(f"OCURRIO UN PROBLEMA {sys.exc_info()[0]}")
      else:
        if nombre_sala == "":
          print("EL DATO NO SE PUEDE OMITIR\n")
        else:
          break

    while True:
      try:
        capacidad = int(input("\nINGRESE LA CAPACIDAD DE LA SALA: "))
      except ValueError:
        print("INGRESE UN VALOR NUMERICO")
      except Exception:
        print(f"OCURRIO UN PROBLEMA {sys.exc_info()[0]}")
      else:
        if capacidad <= 0:
          print("LA CAPACIDAD TIENE QUE SER MAYOR QUE 0")
        else:
          break

    try:
      with sqlite3.connect("registros.db") as conn:
        mi_cursor = conn.cursor()
        valores = {"nombre":nombre_sala, "cupo":capacidad}
        mi_cursor.execute("INSERT INTO sala (nombre, cupo) VALUES(:nombre,:cupo)", valores)
    except Error as e:
      print (e)
    except:
      print(f"Se produjo el siguiente error: {sys.exc_info()[0]}")
    finally:
      conn.close()

    print(f'\n{"*" * 10} SALA REGISTRADA CON EXITO {"*" * 10}')

  elif opcion_menu == 4:
    print(f'\n{"_" * 70}')
    print(f'{" " * 20}  MENU REGISTRO DE UN NUEVO CLIENTE {" " * 20}')
    print(f'{"_" * 70}')

    while True:
      try:
        nombre_cliente = input("INGRESE EL NOMBRE DEL CLIENTE A REGISTRAR: ")
      except Exception:
        print(f"OCURRIO UN PROBLEMA {sys.exc_info()[0]}")
      else:
        if nombre_cliente == "":
          print("EL DATO NO SE PUEDE OMITIR\n")
        else:
          break

    try:
      with sqlite3.connect("registros.db") as conn:
        mi_cursor = conn.cursor()
        valores = {"nombre":nombre_cliente}
        mi_cursor.execute("INSERT INTO cliente (nombre) VALUES(:nombre)", valores)
    except Error as e:
      print (e)
    except:
      print(f"Se produjo el siguiente error: {sys.exc_info()[0]}")
    finally:
      conn.close()

    print(f'\n{"*" * 10} CLIENTE REGISTRADO {"*" * 10}')

  elif opcion_menu == 5:
    break

print("\n¡GRACIAS POR SU PREFERENCIA!")
